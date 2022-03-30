import traceback
from decimal import Decimal

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from internal_api.models import Shop, Supplier, Warehouse
from products.models import MeasurementUnit, Product, ProductUnit


class Command(BaseCommand):
    help = "Loads the product catalog from an Excel file."

    def add_arguments(self, parser):
        parser.add_argument(
            "input_file",
            type=str,
            help="Input file (*.xslx)",
        )
        parser.add_argument(
            "--shop",
            type=int,
            help="Outlet Id to prepare for accepting goods",
        )

    def handle(self, input_file, *args, **options):
        wb = load_workbook(input_file, read_only=True)
        ws = wb["TDSheet"]

        shop_id = options.get("shop", None)
        shop = Shop.objects.get(id=shop_id) if shop_id else None

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                product_name = row[2].value
                if not product_name:
                    continue

                supplier_name = row[1].value
                barcode = int(row[3].value or 0)
                unit_name = row[4].value or "шт"
                cost = Decimal(row[5].value or 0)
                price = Decimal(row[6].value or 0)

                unit, _ = MeasurementUnit.objects.get_or_create(name=unit_name)
                unit.save()

                product, _ = Product.objects.get_or_create(name=product_name)
                product.save()

                product_unit, _ = ProductUnit.objects.get_or_create(
                    product=product,
                    unit=unit,
                    defaults={"barcode": barcode},
                )
                product_unit.save()

                if shop and price and cost:
                    supplier, _ = Supplier.objects.get_or_create(
                        name=supplier_name,
                    )
                    supplier.save()

                    warehouse, _ = Warehouse.objects.get_or_create(
                        shop=shop,
                        product_unit=product_unit,
                        supplier=supplier,
                        defaults={
                            "price": price,
                            "margin": round(price / cost * 100 - Decimal(100), 2),
                        },
                    )
                    warehouse.save()
            except:  # noqa
                print(f"При обработке строки {i} возникло исключение:")
                traceback.print_exc()

        wb.close()
