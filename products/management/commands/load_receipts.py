import traceback
from collections import defaultdict
from decimal import Decimal
from functools import cache

from dateutil import parser
from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils.timezone import get_current_timezone
from openpyxl import load_workbook

from internal_api.models import (
    Batch,
    ReceiptDocument,
    Shop,
    Supplier,
    SupplyContract,
    Warehouse,
    WarehouseOrder,
    WarehouseRecord,
)
from products.models import MeasurementUnit, Product, ProductUnit


@cache
def _get_or_create_shop(address):
    return Shop.objects.get_or_create(
        address=address,
        defaults={"name": f"Филиал на {address[:244]}"},
    )[0]


@cache
def _get_or_create_supply_contract(supplier_id, number, date):
    return SupplyContract.objects.get_or_create(
        supplier_id=supplier_id,
        contract_number=number,
        contract_date=date,
    )[0]


@cache
def _get_or_create_supplier(supplier_name, contract_number, contract_date):
    # multiple suppliers may have the same names,
    # but name+contract must be pretty unique
    supplier = Supplier.objects.filter(
        name=supplier_name,
        supply_contract__contract_number=contract_number,
        supply_contract__contract_date=contract_date,
    ).last()
    if supplier:
        return supplier

    with transaction.atomic():
        supplier = Supplier.objects.create(name=supplier_name)
        SupplyContract.objects.create(
            supplier=supplier,
            contract_number=contract_number,
            contract_date=contract_date,
        )
    return supplier


class Command(BaseCommand):
    help = "Creates an inventory document from an Excel file."

    def add_arguments(self, parser):
        parser.add_argument(
            "input_file",
            type=str,
            help="Input file (*.xlsx)",
        )
        parser.add_argument(
            "--new-batch",
            action="store_true",
            help=(
                "Create new batches for each record"
                " (default: use the latest existing batch)"
            ),
        )

    def handle(self, input_file, new_batch, *args, **options):
        wb = load_workbook(input_file, read_only=True)
        ws = wb["TDSheet"]

        receipts = defaultdict(list)
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            receipts[row[1].value].append(i)

        unit, _ = MeasurementUnit.objects.get_or_create(name="шт")

        for receipt_text, row_indices in receipts.items():
            try:
                receipt_number, receipt_datetime = receipt_text.removeprefix(
                    "Поступление товаров и услуг "
                ).split(" от ")
                receipt_datetime = parser.parse(receipt_datetime).replace(
                    tzinfo=get_current_timezone(),
                )
                first_row = ws[row_indices[0]]

                shop_address = first_row[0].value
                shop = _get_or_create_shop(shop_address)

                supplier_name = first_row[4].value
                contract = first_row[5].value
                contract_number, contract_date = contract.removeprefix(
                    "Договор № "
                ).split(" от ")
                contract_date = parser.parse(contract_date.removesuffix("г.")).date()
                supplier = _get_or_create_supplier(
                    supplier_name,
                    contract_number,
                    contract_date,
                )
                faux_order, _ = WarehouseOrder.objects.get_or_create(
                    supplier=supplier,
                    shop=shop,
                    created_at=receipt_datetime,
                    defaults={"status": "delivered"},
                )
                waybill = first_row[2].value
                receipt, _ = ReceiptDocument.objects.get_or_create(
                    number=receipt_number,
                    defaults={
                        "waybill": waybill,
                        "waybill_date": receipt_datetime.date(),
                        "created_at": receipt_datetime,
                        "order": faux_order,
                    },
                )
                receipt.warehouse_records.all().delete()
            except Exception:  # noqa
                print(f"При обработке {receipt_text} возникло исключение:")
                traceback.print_exc(limit=1, chain=False)
                continue

            for i in row_indices:
                try:
                    row = ws[i]
                    product_name = row[6].value  # a room for improvement
                    barcode = row[7].value
                    barcode = int(barcode) if barcode else None
                    quantity = Decimal(row[8].value)
                    cost = Decimal(row[9].value)
                    price = Decimal(row[10].value)
                    vat_rate = Decimal(row[11].value.rstrip("%"))

                    product, _ = Product.objects.get_or_create(
                        name=product_name,
                        defaults={"vat_rate": vat_rate},
                    )
                    product_unit, _ = ProductUnit.objects.get_or_create(
                        product=product,
                        unit=unit,
                        defaults={"barcode": barcode},
                    )
                    warehouse, _ = Warehouse.objects.get_or_create(
                        shop=shop,
                        product_unit=product_unit,
                        defaults={
                            "price": price,
                            "margin": round(price / cost * 100 - Decimal(100), 2),
                        },
                    )
                    if not new_batch:
                        batch = (
                            Batch.objects.filter(
                                supplier=supplier,
                                warehouse_records__warehouse__product_unit=product_unit,
                            )
                            .order_by("created_at")
                            .last()
                        )
                        if not batch:
                            batch = Batch.objects.create(supplier=supplier)
                    else:
                        batch = Batch.objects.create(supplier=supplier)
                    record = WarehouseRecord(
                        batch=batch,
                        document=receipt,
                        warehouse=warehouse,
                        quantity=quantity,
                        cost=cost,
                    )
                    record.save()
                except Exception:  # noqa
                    print(f"При обработке строки {i} возникло исключение:")
                    traceback.print_exc(limit=1, chain=False)

        wb.close()
